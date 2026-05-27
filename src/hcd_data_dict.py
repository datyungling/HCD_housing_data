"""
HCD APR Data Dictionary Consolidator
-------------------------------------
Downloads all 11 HCD APR data dictionary DOCX files and consolidates
them into a single Excel workbook with:
  - One sheet per table (Tab A, A2, C, D, E, F, F2, G, H, I, K)
  - A summary "Overview" sheet describing each table
  - Consistent formatting throughout

Run: python build_hcd_data_dictionary.py
Output: hcd_apr_data_dictionary.xlsx
"""

import os
import re
import requests
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Table metadata ────────────────────────────────────────────────────────────

TABLES = [
    {
        "name": "Table A",
        "sheet": "Table A",
        "resource_id": "198bded5-7a17-4e1e-bea5-688dbb439c6d",
        "filename": "annualprogressreport_tablea_datadictionary-1.docx",
        "description": "Annual housing unit permit activity by jurisdiction. Tracks permits issued, "
                       "units entitled/permitted/completed by affordability level. PRIMARY table "
                       "for building permit validation against Shovels data.",
        "relevance": "HIGH – directly maps to building permit counts",
    },
    {
        "name": "Table A2",
        "sheet": "Table A2",
        "resource_id": "354a375a-821d-4d4e-97ab-4d79c2925fce",
        "filename": "annualprogressreport_tablea2_datadictionary-2.docx",
        "description": "Annual Building Activity Report Summary – New construction, alterations, "
                       "demolitions by unit type. Supplements Table A with structural detail.",
        "relevance": "HIGH – unit type breakdown useful for cross-validation",
    },
    {
        "name": "Table C",
        "sheet": "Table C",
        "resource_id": "2d5e0b66-8d0e-46d5-a32f-6a2f43fdf941",
        "filename": "annualprogressreport_tablec_datadictionary.docx",
        "description": "Sites identified or rezoned to accommodate RHNA (Regional Housing Needs "
                       "Allocation). Tracks zoning changes and site capacity.",
        "relevance": "LOW – zoning/entitlement data, not permit activity",
    },
    {
        "name": "Table D",
        "sheet": "Table D",
        "resource_id": "53a73ff2-d2dc-4590-8629-c4e06808e217",
        "filename": "annualprogressreport_tabled_datadictionary.docx",
        "description": "Program implementation status. Tracks housing programs from the Housing "
                       "Element, their objectives, status, and responsible agencies.",
        "relevance": "LOW – policy/program tracking, not permit activity",
    },
    {
        "name": "Table E",
        "sheet": "Table E",
        "resource_id": "1c80d799-2770-4c0e-bbcf-dd47f9a49496",
        "filename": "annualprogressreport_tablee_datadictionary.docx",
        "description": "Commercial development bonus. Tracks commercial projects that triggered "
                       "residential affordable housing requirements.",
        "relevance": "LOW – commercial development linkage fees",
    },
    {
        "name": "Table F",
        "sheet": "Table F",
        "resource_id": "684f6752-9755-4be0-b5dc-433d30b3203d",
        "filename": "annualprogressreport_tablef_datadictionary.docx",
        "description": "Units rehabilitated, preserved, or acquired with public assistance. "
                       "Tracks affordable housing preservation activity.",
        "relevance": "MEDIUM – rehab permits may overlap with Shovels data",
    },
    {
        "name": "Table F2",
        "sheet": "Table F2",
        "resource_id": "01c77ab2-c095-4b7b-bed7-86a6e29bdfe7",
        "filename": "apr-table-f2-data-dictionary.docx",
        "description": "Moderate-income housing (up to 25% of jurisdiction RHNA). Tracks "
                       "moderate-income units counted toward RHNA obligations.",
        "relevance": "LOW – affordability tracking, limited permit overlap",
    },
    {
        "name": "Table G",
        "sheet": "Table G",
        "resource_id": "0ddb8526-c758-4d73-ae8c-d72125a2aea5",
        "filename": "annualprogressreport_tableg_datadictionary.docx",
        "description": "Locally owned lands declared surplus. Tracks surplus government land "
                       "available for affordable housing development.",
        "relevance": "LOW – land inventory, not permit activity",
    },
    {
        "name": "Table H",
        "sheet": "Table H",
        "resource_id": "0499a106-5667-437e-acab-362c44283bb1",
        "filename": "annualprogressreport_tableh_datadictionary.docx",
        "description": "RHNA progress summary. Aggregate count of units permitted vs. RHNA "
                       "targets by income category for the planning period.",
        "relevance": "MEDIUM – aggregate permit counts useful for trend validation",
    },
    {
        "name": "Table I",
        "sheet": "Table I",
        "resource_id": "3e9406e6-b854-4a5f-b558-7d15078b8a36",
        "filename": "apr-table-i-data-dictionary.docx",
        "description": "Units demolished or converted. Tracks loss of housing units through "
                       "demolition, conversion, or other removal from housing stock.",
        "relevance": "MEDIUM – demolition permits may appear in Shovels data",
    },
    {
        "name": "Table K",
        "sheet": "Table K",
        "resource_id": "a710f7e7-e469-4c69-9e6d-b2b71e4f8460",
        "filename": "apr-table-k-data-dictionary.docx",
        "description": "Local tenant protection policies. Tracks just-cause eviction ordinances, "
                       "rent control, and other tenant protections by jurisdiction.",
        "relevance": "LOW – policy tracking, no permit activity",
    },
]

BASE_DOWNLOAD_URL = (
    "https://data.ca.gov/dataset/81b0841f-2802-403e-b48e-2ef4b751f77c/resource"
    "/{resource_id}/download/{filename}"
)

# ── DOCX parser ───────────────────────────────────────────────────────────────

WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"

def extract_text_from_cell(cell_el):
    texts = []
    for t in cell_el.iter(f"{{{WORD_NS}}}t"):
        if t.text:
            texts.append(t.text)
    return " ".join(texts).strip()

def parse_docx_tables(docx_path):
    """Extract all tables from a DOCX file as list of list-of-rows."""
    results = []
    with zipfile.ZipFile(docx_path, "r") as z:
        with z.open("word/document.xml") as f:
            tree = ET.parse(f)
    root = tree.getroot()
    for tbl in root.iter(f"{{{WORD_NS}}}tbl"):
        rows = []
        for tr in tbl.iter(f"{{{WORD_NS}}}tr"):
            row = [extract_text_from_cell(tc) for tc in tr.iter(f"{{{WORD_NS}}}tc")]
            if any(row):
                rows.append(row)
        if rows:
            results.append(rows)
    return results

def parse_docx_paragraphs(docx_path):
    """Extract plain text paragraphs (fallback if no tables found)."""
    paragraphs = []
    with zipfile.ZipFile(docx_path, "r") as z:
        with z.open("word/document.xml") as f:
            tree = ET.parse(f)
    root = tree.getroot()
    for para in root.iter(f"{{{WORD_NS}}}p"):
        texts = [t.text for t in para.iter(f"{{{WORD_NS}}}t") if t.text]
        line = "".join(texts).strip()
        if line:
            paragraphs.append(line)
    return paragraphs

# ── Style helpers ─────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL   = PatternFill("solid", fgColor="2E75B6")
ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")
HIGH_FILL     = PatternFill("solid", fgColor="C6EFCE")
MED_FILL      = PatternFill("solid", fgColor="FFEB9C")
LOW_FILL      = PatternFill("solid", fgColor="F4CCCC")

WHITE_BOLD    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
WHITE_NORMAL  = Font(name="Arial", color="FFFFFF", size=10)
BLACK_BOLD    = Font(name="Arial", bold=True, size=10)
BLACK_NORMAL  = Font(name="Arial", size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

def style_header(cell, dark=True):
    cell.fill   = HEADER_FILL if dark else SUBHDR_FILL
    cell.font   = WHITE_BOLD
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER

def style_data(cell, alt=False, bold=False):
    if alt:
        cell.fill = ALT_FILL
    cell.font   = BLACK_BOLD if bold else BLACK_NORMAL
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = THIN_BORDER

def relevance_fill(relevance_str):
    if "HIGH" in relevance_str:
        return HIGH_FILL
    if "MEDIUM" in relevance_str:
        return MED_FILL
    return LOW_FILL

# ── Sheet builders ────────────────────────────────────────────────────────────

def build_overview_sheet(ws, tables):
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "HCD APR Data Dictionary – Table Overview"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "Source: California Department of Housing and Community Development (HCD) | "
        "data.ca.gov | Updated weekly"
    )
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Column headers
    headers = ["Table", "Sheet Name", "Description", "Relevance to Shovels Validation",
               "Resource ID", "Direct Download URL"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        style_header(c)
    ws.row_dimensions[4].height = 20

    for i, t in enumerate(tables):
        row = 5 + i
        url = BASE_DOWNLOAD_URL.format(resource_id=t["resource_id"], filename=t["filename"])
        values = [t["name"], t["sheet"], t["description"], t["relevance"],
                  t["resource_id"], url]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            style_data(c, alt=(i % 2 == 1))
            if col == 4:
                c.fill = relevance_fill(t["relevance"])
                c.font = BLACK_BOLD

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 55

    for row in range(5, 5 + len(tables)):
        ws.row_dimensions[row].height = 45

    # Legend
    legend_row = 5 + len(tables) + 2
    ws.cell(row=legend_row, column=1, value="Relevance Legend:").font = BLACK_BOLD
    items = [("HIGH", HIGH_FILL), ("MEDIUM", MED_FILL), ("LOW", LOW_FILL)]
    for j, (label, fill) in enumerate(items):
        c = ws.cell(row=legend_row, column=2 + j, value=label)
        c.fill = fill
        c.font = BLACK_BOLD
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER


def build_table_sheet(ws, table_meta, docx_path):
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:E1")
    ws["A1"] = f"HCD APR – {table_meta['name']} Data Dictionary"
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Description row
    ws.merge_cells("A2:E2")
    ws["A2"] = table_meta["description"]
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Relevance row
    ws.merge_cells("A3:E3")
    ws["A3"] = f"Relevance to Shovels validation: {table_meta['relevance']}"
    rel_c = ws["A3"]
    rel_c.fill = relevance_fill(table_meta["relevance"])
    rel_c.font = BLACK_BOLD
    rel_c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 18

    # Parse DOCX
    tables_found = parse_docx_tables(docx_path)

    data_start_row = 5
    if tables_found:
        # Use first table; treat first row as header
        tbl = tables_found[0]
        header_row = tbl[0]
        # Normalize: ensure at least 3 cols
        max_cols = max(len(r) for r in tbl)

        for col, h in enumerate(header_row, 1):
            c = ws.cell(row=data_start_row, column=col, value=h)
            style_header(c)
        ws.row_dimensions[data_start_row].height = 20

        for i, row in enumerate(tbl[1:], 1):
            excel_row = data_start_row + i
            for col in range(1, max_cols + 1):
                val = row[col - 1] if col <= len(row) else ""
                c = ws.cell(row=excel_row, column=col, value=val)
                style_data(c, alt=(i % 2 == 0), bold=(col == 1))
            ws.row_dimensions[excel_row].height = 40

        # Column widths
        col_widths = [30, 15, 55, 20, 20]
        for ci, w in enumerate(col_widths[:max_cols], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    else:
        # Fallback: paragraph text
        paras = parse_docx_paragraphs(docx_path)
        ws.cell(row=data_start_row, column=1, value="Field / Content").font = BLACK_BOLD
        for i, p in enumerate(paras):
            c = ws.cell(row=data_start_row + 1 + i, column=1, value=p)
            style_data(c, alt=(i % 2 == 0))
            ws.row_dimensions[data_start_row + 1 + i].height = 30
        ws.column_dimensions["A"].width = 100

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    download_dir = "hcd_dicts"
    os.makedirs(download_dir, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Build overview sheet first
    overview_ws = wb.create_sheet("Overview")
    build_overview_sheet(overview_ws, TABLES)

    for t in TABLES:
        docx_path = os.path.join(download_dir, t["filename"])

        # Download if not already cached
        if not os.path.exists(docx_path):
            url = BASE_DOWNLOAD_URL.format(
                resource_id=t["resource_id"], filename=t["filename"]
            )
            print(f"Downloading {t['name']}...")
            r = requests.get(url, timeout=30)
            r.raise_for_status()
            with open(docx_path, "wb") as f:
                f.write(r.content)
            print(f"  → {len(r.content):,} bytes")
        else:
            print(f"Using cached {t['name']}")

        ws = wb.create_sheet(t["sheet"])
        build_table_sheet(ws, t, docx_path)
        print(f"  → Sheet '{t['sheet']}' built")

    output = "hcd_apr_data_dictionary.xlsx"
    wb.save(output)
    print(f"\nDone! Saved: {output}")

if __name__ == "__main__":
    main()